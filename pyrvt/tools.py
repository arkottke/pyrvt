#!/usr/bin/python
# -*- coding: utf-8 -*-

"""
Tools for reading/writing of files and performing operations.
"""

import csv
import functools
import multiprocessing
import glob
import os
import sys

import numpy as np
import pyprind

from pyrvt.peak_calculators import get_peak_calculator, get_region
from pyrvt import motions

# Try to load the modules required for reading/writing files
try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import xlwt
except ImportError:
    xlwt = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

PARAMETER_NAMES = [
    ('magnitude', 'Magnitude'),
    ('distance', 'Distance (km)'),
    ('vs30', 'Vs30 (m/s)'),
    ('kappa', 'Site Atten., Kappa0 (sec)'),  # Site Atten., κ₀
    ('duration', 'Duration (sec)'),
    ('region', 'Region'),
]


def read_events(fname, response_type):
    """Read data from the file an Excel work book.

    Parameters
    ----------
    fname : str
        Filename of the input file.
    response_type : str
        Type of response. Valid options are: 'psa' for psuedo-spectral
        acceleration, or 'fa' for Fourier amplitude.

    Returns
    -------
    ext : str
        Extension of input file
    reference : :class:`numpy.ndarray`
        Reference of the response. This is either period (sec) for
        response_type 'psa' or frequency (Hz) for response_type 'fa'
    events : List[dict]
        List of events read from the file. See ``Note`` in
        :func:`.calc_compatible_spectra` for more information on structure of
        the dictionaries.
    """
    assert response_type in ['psa', 'fa']

    ext = os.path.splitext(fname)[1].lower()
    # Load the file depending on the format
    if ext == '.csv':
        def parse(s):
            try:
                return float(s)
            except ValueError:
                return s

        with open(fname) as fp:
            reader = csv.reader(fp)
            rows = [[parse(r) for r in row] for row in reader]
    elif ext == '.xls':
        if xlrd is None:
            raise RuntimeError('xlrd is required to open an xls file')
        wb = xlrd.open_workbook(fname)
        ws = wb.sheet_by_index(0)
        rows = [ws.row_values(i) for i in range(ws.nrows)]
    elif ext == '.xlsx':
        if openpyxl is None:
            raise RuntimeError('openpyxl is required to open an xlsx file')
        wb = openpyxl.load_workbook(fname, read_only=True)
        ws = wb.worksheets[0]
        rows = [[r.value for r in row] for row in ws.rows]
        # Close the file so that it may be deleted if needed. This is only
        # important so that in the test cases the temporary .xlsx file can be
        # deleted.
        wb._archive.close()
    else:
        raise NotImplementedError

    parameters = {key: rows[i][1:]
                  for i, (key, label) in enumerate(PARAMETER_NAMES)}

    event_row = len(parameters) + 1
    event_count = len(rows[0]) - 1

    reference = np.array([row[0] for row in rows[event_row:]])

    events = []
    for i in range(event_count):
        resps = np.array([row[i + 1] for row in rows[event_row:]])
        # Extract the appropriate attributes
        e = {k: v[i] for k, v in parameters.items()}
        e[response_type] = resps

        if 'region' in e:
            e['region'] = get_region(e['region'])

        events.append(e)

    return ext, reference, events


def write_events(fname, reference, reference_label, response_type,
                 response_label, events):
    """Write the events to a file.

    Parameters
    ----------
    fname : str
        Save the events to this file. The directory is created if needed.
    reference : array_like
        Periods of the oscillator response shared across all events.
    reference_label : str
        Label of the reference (e.g., 'Frequency (Hz)').
    response_type : str
        Type of response. Valid options: `psa` for pseudo-spectral
        accleration, or `fa` for Fourier amplitude.
    response_label : str
        Label of the response type (e.g., 'Fourier Ampl. (g/sec)')
    events : List[dict]
        Events to write to file. See ``Note`` in
        :func:`.compute_compatible_spectra` for more information.

    Raises
    ------
    NotImplementedError:
        If extension is not supported
    """
    # Create the rows of output
    rows = []
    # Output the parameters
    for key, label in PARAMETER_NAMES:
        rows.append([label] + [e[key] for e in events])

    rows.append(
        [reference_label] + len(events) * [response_label])

    # Output the response spectra
    for i in range(len(reference)):
        rows.append(
            [reference[i]] + [e[response_type][i] for e in events])

    # Create the directory
    dirname = os.path.dirname(fname)
    if not os.path.exists(dirname):
        os.makedirs(dirname)

    # Write the file
    ext = os.path.splitext(fname)[1].lower()

    if ext == '.csv':
        if sys.version_info < (3, 1):
            fp = open(fname, 'wt')
        else:
            fp = open(fname, 'wt', newline='')
        writer = csv.writer(fp)
        writer.writerows(rows)
        fp.close()
    elif ext == '.xls':
        if xlwt is None:
            raise RuntimeError('xlwt is required to open an xls file')
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Sheet 1')
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                ws.write(i, j, cell)
        wb.save(fname)
    elif ext == '.xlsx':
        if openpyxl is None:
            raise RuntimeError('openpyxl is required to open an xlsx file')
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        for row in rows:
            ws.append(row)
        wb.save(fname)
    else:
        raise NotImplementedError


def _calc_fa(bar, target_freqs, damping, method, event):
    """Calculate the fourier amplitudes for an event.
    
    Note that this is intended as a helper function to be called by 
    multiprocessing.Pool.
    """
    event_keys = ['magnitude', 'distance', 'region']
    event_kwds = {key: event[key] for key in event_keys}
    crm = motions.CompatibleRvtMotion(
        target_freqs,
        event['psa'],
        duration=event['duration'],
        osc_damping=damping,
        event_kwds=event_kwds,
        peak_calculator=get_peak_calculator(
            method, event_kwds)
    )
    psa_calc = crm.calc_osc_accels(target_freqs, damping)
    if bar:
        bar.update()
    return crm, psa_calc


def calc_compatible_spectra(method, periods, events, damping=0.05,
                            verbose=True):
    """Compute the response spectrum compatible motions.

    Parameters
    ----------
    method : str
        RVT peak factor method, see
        :func:`~.peak_calculators.get_peak_calculator`.
    periods : array_like
        Periods of the oscillator response shared across all events.
    events : List[dict]
        All events to consider. See ``Note``.
    damping : float, optional
        Fractional damping of the oscillator (decimal). Default value of 0.05
        for a damping ratio of 5%.
    verbose : bool, optional
        Print status of calculation. Default is `True`.

    Returns
    -------
    :class:`numpy.ndarray`
        Frequency of the computed Fourier amplitude spectra.

    Note
    ----
    Each event dictionary should have the following keys:

    - **psa** : :class:`numpy.ndarray` -- pseudo-spectral accelerations. This
      is the target for the :class:`~.motions.CompatibleRvtMotion`.
    - **duration** : float, optional -- duration of the ground motion
    - **magnitude** : float, optional -- earthquake magnitude
    - **distance** : float, optional -- earthquake distance (km)
    - **region** : str, optional -- earthquake source region, see
      :func:`~.peak_calculators.get_region` If no duration is provided one
      is estimated from the magnitude, distance, and region.

    The `events` dictionary is modified by this function and adds the
    following keys:

    - **duration** : float -- duration of the ground motion if one was not
      specified
    - **fa** : :class:`numpy.ndarray` -- Fourier amplitude spectra in units of
      g/sec
    - **psa_calc** : :class:`numpy.ndarray` -- Pseudo-spectral acceleration
      calculated from `fa`. This will differ slightly from `psa_target`.
    """
    bar = pyprind.ProgPercent(len(events)) if verbose else None
    target_freqs = 1. / periods
    with multiprocessing.Pool() as pool:
        results = pool.map(
            functools.partial(_calc_fa, bar, target_freqs, damping, method),
            events)

    # Copy values back into the dictionary
    for event, (crm, psa_calc) in zip(events, results):
        if not event['duration']:
            event['duration'] = crm.duration
        event['fa'] = crm.fourier_amps
        event['psa_calc'] = psa_calc

    # Return the frequency from one of the computed motions.
    freqs = results[0][0].freqs
    return freqs


def operation_psa2fa(src, dst, damping, method='LP99', fixed_spacing=True,
                     verbose=True):
    """Compute the acceleration response spectrum from a Fourier amplitude
    spectrum.

    Parameters
    ----------
    src : str
        Source for the pseudo-spectral accelerations (PSA). This can be a
        filename or pattern used in :func:`glob.glob`.
    dst : str
        Destination directory for the output PSA. The directory is created if
        it does not exist.
    damping : float
        Fractional damping of the oscillator ( decimal).
    method : str
        RVT peak factor method, see
        :func:`~.peak_calculators.get_peak_calculator`.
    fixed_spacing : bool, optional
        If `True`, then the periods are interpolated to 301 points equally
        space in log-space from 0.01 to 10.
    verbose : bool, optional
        Print status of calculation.
    """
    for filename_src in glob.iglob(src):
        if verbose:
            print('Processing:', filename_src)
        ext, periods, events = read_events(filename_src, 'psa')

        if fixed_spacing:
            # Interpolate the periods to a smaller range
            _periods = np.logspace(-2, 1, 301)

            for e in events:
                e['psa'] = np.exp(np.interp(
                    _periods, periods, np.log(e['psa'])))

            periods = _periods

        # Compute the FA from the PSA
        freqs = calc_compatible_spectra(method, periods, events,
                                        damping=damping, verbose=verbose)

        if not os.path.exists(dst):
            os.makedirs(dst)

        basename = os.path.basename(filename_src)
        pathname_dst = os.path.join(dst, basename.rsplit('_', 1)[0])

        write_events(pathname_dst + '_sa' + ext, periods, 'Period (s)',
                     'psa_calc', 'Sa (g)', events)
        write_events(pathname_dst + '_fa' + ext, freqs, 'Frequency (Hz)',
                     'fa', 'FA (g-s)', events)


def _calc_psa(bar, osc_freqs, damping, method, freqs, event):
    """Calculate the response spectra for an event.

    Note that this is intended as a helper function to be called by 
    multiprocessing.Pool.
    """
    m = motions.RvtMotion(
        freqs=freqs,
        fourier_amps=event['fa'],
        duration=event['duration'],
        peak_calculator=get_peak_calculator(
            method, dict(region=event['region'], mag=event['magnitude'],
                         dist=event['distance']))
    )
    psa = m.calc_osc_accels(osc_freqs, damping)
    if bar:
        bar.update()
    return psa


def operation_fa2psa(src, dst, damping, method='LP99', fixed_spacing=True,
                     verbose=True):
    """Compute the Fourier amplitude spectrum from a acceleration response
    spectrum.

    Parameters
    ----------
    src : str
        Source for the Fourier amplitudes. This can be a filename or pattern
        used in :func:`glob.glob`.
    dst : str
        Destination directory for the output PSA. The directory is created if
        it does not exist.
    damping : float
        Fractional damping of the oscillator (decimal).
    method : str
        RVT peak factor method, see
        :func:`~.peak_calculators.get_peak_calculator`.
    fixed_spacing : bool, optional
        If `True`, then the periods are interpolated to 301 points equally
        space in log-space from 0.01 to 10.
    verbose : bool, optional
        Print status of calculation.
    """
    if fixed_spacing:
        periods = np.logspace(-2, 1, 301)
        osc_freqs = 1. / periods

    for filename_src in glob.iglob(src):
        if verbose:
            print('Processing:', filename_src)
        ext, freqs, events = read_events(filename_src, 'fa')

        if not fixed_spacing:
            osc_freqs = freqs
            periods = 1. / osc_freqs

        bar = pyprind.ProgPercent(len(events)) if verbose else None
        with multiprocessing.Pool() as pool:
            psas = pool.map(
                functools.partial(_calc_psa, bar, osc_freqs, damping, method, freqs),
                events
            )

        for event, psa in zip(events, psas):
            event['psa'] = psa

        if not os.path.exists(dst):
            os.makedirs(dst)

        basename = os.path.basename(filename_src)

        pathname_dst = os.path.join(dst, basename.rsplit('_', 1)[0])

        write_events(pathname_dst + '_sa' + ext, periods, 'Period (s)', 'psa',
                     'PSA (g)', events)
