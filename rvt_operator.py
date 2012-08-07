#!/usr/bin/python

# This file is part of IRVT.
#
# Foobar is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# IRVT is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with IRVT.  If not, see <http://www.gnu.org/licenses/>.

'''
File: rvt_operator.py
Author: Albert Kottke
Description: Provides a command line iterface for performing RVT calculations.
'''


import argparse
import csv
import glob
import os

import numpy as np

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import xlwt
except ImportError:
    xlrd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

import rvt

damping = 0.05
parameter_names = [
    ('magnitude', 'Magnitude'),
    ('distance', 'Distance (km)'),
    ('vs30', 'Vs30 (m/s)'),
    ('kappa', 'Kappa0 (sec)'),
    ('duration', 'Duration (sec)'),
        ]

def load_events(filename, response_key='sa'):
    '''Read data from the file an Excel work book'''

    ext = os.path.splitext(filename)[1].lower()
    # Load the file depending on the format
    if ext == '.csv':
        def parse(s):
            try:
                return float(s)
            except ValueError:
                return s

        with open(filename) as fp:
            reader = csv.reader(fp)
            rows = [[parse(r) for r in row] for row in reader]
    elif ext == '.xls':
        if xlrd is None:
            raise RuntimeError, 'xlrd is required to open an xls file'

        wb = xlrd.open_workbook(filename)
        ws = wb.sheet_by_index(0)
        rows = [ws.row_values(i) for i in range(ws.nrows)]

    elif ext == '.xlsx':
        if openpyxl is None:
            raise RuntimeError, 'openpyxl is required to open an xlsx file'

        # FIXME Not found?
        wb = openpyxl.load_workbook(filename)
        ws = wb.worksheets[0]
        rows = [[r.value for r in row] for row in ws.rows]
    else:
        raise NotImplementedError

    parameters = {key: rows[i][1:]
            for i, (key, label) in enumerate(parameter_names)}

    event_row = len(parameters) + 1
    event_count = len(rows[0]) - 1

    reference = np.array([row[0] for row in rows[event_row:]])

    events = []
    for i in range(event_count):
        sa = np.array([row[i + 1] for row in rows[event_row:]])
        # Extract the appropriate attributes
        e = {k: v[i] for k, v in parameters.iteritems()}
        e[response_key] = sa

        if e['duration'] in ['wus', 'ceus']:
            e['region'] = e['duration']
            e['duration'] = None
        else:
            e['region'] = None

        events.append(e)

    return ext, reference, events


def export_events(filename, reference, reference_label, response_key,
        response_label, events):

    # Create the rows of output
    rows = []
    # Output the parameters
    for key, label in parameter_names:
        rows.append([label] + [e[key] for e in events])

    rows.append(
            [reference_label] + len(events) * [response_label])

    # Output the response spectra
    for i in range(len(reference)):
        rows.append(
            [reference[i]] + [e[response_key][i] for e in events])

    # Create the directory
    dirname = os.path.dirname(filename)

    if not os.path.exists(dirname):
        os.makedirs(dirname)

    # Write the file
    ext = os.path.splitext(filename)[1].lower()

    if ext == '.csv':
        with open(filename, 'wb') as fp:
            writer = csv.writer(fp)
            writer.writerows(rows)
    elif ext == '.xls':
        if xlwt is None:
            raise RuntimeError, 'xlwt is required to open an xls file'

        wb = xlwt.Workbook()
        ws = wb.add_sheet('Sheet 1')

        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                ws.write(i, j, cell)

        wb.save(filename)
    elif ext == '.xlsx':
        if openpyxl is None:
            raise RuntimeError, 'openpyxl is required to open an xlsx file'

        wb = openpyxl.Workbook()
        ws = wb.add_sheet('Sheet 1')

        map(ws.append, rows)

        wb.save(filename)
    else:
        raise NotImplementedError

def compute_compatible_spectra(period, events, damping=0.05):
    ''' Compute the response spectrum compatible motions. '''
    target_freq = 1. / period

    for e in events:
        crm = rvt.CompatibleRvtMotion(target_freq, e['sa_target'],
                damping=damping, magnitude=e['magnitude'],
                distance=e['distance'], duration=e['duration'],
                region=e['region'])

        freq = crm.freq

        if not e['duration']:
            e['duration'] = crm.duration

        e['fa'] = crm.fourier_amp
        e['sa_calc'] = crm.compute_osc_resp(target_freq, damping)

    return freq


def operation_sa2fa(src, dest, damping, fixed_spacing):
    '''Compute the acceleration response spectrum from a Fourier amplitude
    spectrum.'''

    for filename_src in glob.iglob(src):
        ext, period, events = load_events(filename_src, 'sa_target')

        if fixed_spacing:
            # Interpolate the periods to a smaller range
            _period = np.logspace(-2, 1, 100)

            for e in events:
                e['sa_target'] = np.exp(np.interp(
                    _period, period, np.log(e['sa_target'])))

            period = _period

        # Compute the FA from the PSA
        freq = compute_compatible_spectra(period, events, damping=damping)

        if not os.path.exists(dest):
            os.makedirs(dest)

        basename = os.path.basename(filename_src)
        pathname_dest = os.path.join(dest, basename.rsplit('_', 1)[0])

        export_events(pathname_dest + '_sa' + ext, period, 'Period (s)',
                'sa_calc', 'Sa (g)', events)
        export_events(pathname_dest + '_fa' + ext, freq, 'Frequency (Hz)', 'fa',
                'FA (g-s)', events)


def operation_fa2sa(src, dest, damping, fixed_spacing):
    '''Compute the Fourier amplitude spectrum from a acceleration response
    spectrum.'''

    if fixed_spacing:
        period = np.logspace(-2, 1, 100)
        osc_freq = 1. / period

    for filename_src in glob.iglob(src):
        ext, freq, events = load_events(filename_src, 'fa')

        if not fixed_spacing:
            osc_freq = freq
            period = 1. / osc_freq

        for e in events:
            m = rvt.RvtMotion(freq=freq, fourier_amp=e['fa'],
                    duration=e['duration'])
            e['sa'] = m.compute_osc_resp(osc_freq, damping)

        if not os.path.exists(dest):
            os.makedirs(dest)

        basename = os.path.basename(filename_src)

        pathname_dest = os.path.join(dest, basename.rsplit('_', 1)[0])

        export_events(pathname_dest + '_sa' + ext, period, 'Period (s)', 'sa',
                'Sa (g)', events)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Compute reponse or Fourier amplitude spectra using RVT.')
    parser.add_argument('operation',
            help='''Operation to be performed. [sa2fa] converts from
            (psuedo)-spectral acceleration to Fourier amplitude.  [fa2sa]
            converts from Fourier amplitude to (psuedo)-spectral acceleration.
            ''',
            choices=['sa2fa', 'fa2sa'])
    parser.add_argument('-i', '--input', dest='src',
            help='''Path containing the input file(s). Supported file types
            are csv, xls, and xlsx -- provided the required packages have been
            installed. A single file or glob can be specified. An example of a
            glob would be "input/*_sa.xls" for all files within directory
            "input" ending in "_sa.xls".''',
            required=True)
    parser.add_argument('-o', '--output', dest='dest',
            help='''Path where the output files should be created. If this
            directory does not exist it will be created. Default: ./output''',
            default='./output')
    parser.add_argument('-d', '--damping', dest='damping', default=0.05,
            help='''Oscillator damping in decimal.  Default: 0.05.''')
    parser.add_argument('-f', '--fixed-spacing', dest='fixed_spacing',
            action='store_true',
            help='''Fixed spacing of the oscillator period of 0.01 to 10 sec
            log-spaced with 100 points. Target SA values will be interpolated
            if needed''')

    args = parser.parse_args()

    if args.operation == 'sa2fa':
        operation_sa2fa(args.src, args.dest, args.damping, args.fixed_spacing)
    elif args.operation == 'fa2sa':
        operation_fa2sa(args.src, args.dest, args.damping, args.fixed_spacing)
    else:
        raise NotImplementedError
